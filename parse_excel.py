import openpyxl
import sys

# Открываем файл Excel
wb = openpyxl.load_workbook(r'c:\Users\miros\Downloads\Stock Tracker (4).xlsx')

# Выводим названия всех листов
print("📋 Листы в файле:")
for sheet_name in wb.sheetnames:
    print(f"   - {sheet_name}")

# Берем первый (активный) лист
ws = wb.active
print(f"\n📊 Анализ листа: {ws.title}")
print("="*80)

# Читаем заголовки (первая строка)
headers = []
for cell in ws[1]:
    headers.append(cell.value if cell.value else "")

print(f"\n📝 Заголовки ({len(headers)} колонок):")
for i, header in enumerate(headers, 1):
    if header:
        print(f"   {i}. {header}")

# Читаем первые 10 строк данных
print(f"\n📦 Первые 5 строк данных:")
print("-"*80)

for row_idx in range(2, min(7, ws.max_row + 1)):
    print(f"\nСтрока {row_idx}:")
    row_data = []
    for col_idx, cell in enumerate(ws[row_idx], 1):
        if col_idx <= len(headers) and headers[col_idx-1]:
            value = cell.value if cell.value is not None else ""
            print(f"   {headers[col_idx-1]}: {value}")

print(f"\n📊 Итого строк в файле: {ws.max_row}")
print(f"📊 Итого колонок: {ws.max_column}")

# Проверяем, есть ли артикул Its1_2_3/50g
print(f"\n🔍 Поиск артикула Its1_2_3/50g:")
found = False
for row_idx in range(2, ws.max_row + 1):
    first_cell = ws.cell(row=row_idx, column=1).value
    if first_cell and "Its1_2_3/50g" in str(first_cell):
        found = True
        print(f"   ✅ Найден в строке {row_idx}")
        print(f"   Данные строки:")
        for col_idx in range(1, min(len(headers) + 1, ws.max_column + 1)):
            if col_idx <= len(headers) and headers[col_idx-1]:
                value = ws.cell(row=row_idx, column=col_idx).value
                print(f"      {headers[col_idx-1]}: {value}")
        break

if not found:
    print("   ❌ Артикул Its1_2_3/50g не найден")

wb.close()
